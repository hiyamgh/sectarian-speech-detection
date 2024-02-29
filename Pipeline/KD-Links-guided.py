import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import itertools
import os
from selenium.webdriver.chrome.options import Options
import concurrent.futures as futures
import random
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import re
from tqdm import tqdm

TWITTER_USERNAME = None
TWITTER_PASSWORD = None


def timeout(timelimit):
    def decorator(func):
        def decorated(*args, **kwargs):
            with futures.ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(func, *args, **kwargs)
                try:
                    result = future.result(timelimit)
                except futures.TimeoutError:
                    print('Timedout!')
                    raise TimeoutError from None
                else:
                    print(result)
                executor._threads.clear()
                futures.thread._threads_queues.clear()
                return result
        return decorated
    return decorator


def mkdir(folder_name):
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)


def line_split(line):
    return re.findall(r'[^"\s]\S*|".+?"', line)


def get_accounts_per_entity():
    ''' creates a mapping between an entity (channel) and social media accounts associated with that entity '''
    df = pd.read_excel('Accounts.xlsx')
    entities = list(df['Name'])
    facebook_accs = list(df['Facebook Account'])
    twitter_accs = list(df['Twitter Account '])
    instagram_accs = list(df['Instagram Account'])
    youtube_accs = list(df['Youtube Account'])

    entities2accs = {} # dictionary mapping entities to accounts
    for i, ent in enumerate(entities):
        entities2accs[ent] = {}
        entities2accs[ent]['facebook'] = facebook_accs[i]
        entities2accs[ent]['twitter'] = twitter_accs[i]
        entities2accs[ent]['instagram'] = instagram_accs[i]
        entities2accs[ent]['youtube'] = youtube_accs[i]

    return entities2accs


def get_youtube_volumes():
    videos = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'ytd-video-renderer')))
    all_links = []
    meta_data = []
    days = []
    all_comments_volume = []

    for v in videos:
        meta = v.find_element_by_css_selector('div#metadata').text
        link = v.find_element_by_css_selector('a#video-title').get_attribute('href')
        all_links.append(link)
        meta_data.append(meta)
        print('meta: {}\nvideo link: {}\n\n'.format(meta, link))

    print('=========================================================================')
    for link in all_links:
        driver.get(link)
        time.sleep(random.randint(5, 10))
        snippet = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div#snippet')))
        driver.execute_script("arguments[0].scrollIntoView();", snippet[0])
        driver.execute_script("arguments[0].click();", snippet[0])
        day = driver.find_element_by_css_selector('ytd-watch-info-text#ytd-watch-info-text').text
        print(day.strip())

        day_list = day.strip().split(' ')
        day_actual = day.strip()
        for i in range(len(day_list)):
            if day_list[i] in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
                day_actual = ' '.join(day_list[i-1:i+2])
                break
        print('Day Actual: {}'.format(day_actual))
        days.append(day_actual)

        try:
            comments = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div#title h2#count')))
            # comments_list = comments.text.strip().split(' ')
            # comments_volume = comments.text.strip()
            # for i in range(len(comments_list)):
            #     if comments_list[i] == 'Comments':
            #         comments_volume = ' '.join(comments_list[i-1:i+1])
            #         break
            # print('Comments volume: {}'.format(comments_volume))
            # all_comments_volume.append(comments_volume)
            print('Comments volume: {}'.format(comments.text))
            all_comments_volume.append(comments.text)
        except:
            print('Comments are turned off. Learn more')
            all_comments_volume.append('Comments are turned off. Learn more')

        if len(all_comments_volume) % 5 == 0:
            data = {
                'Link': all_links,
                'Meta Data': meta_data,
                'Day': days + ['' for _ in range(len(all_links) - len(days))],
                'Comments Volume': all_comments_volume + ['' for _ in range(len(all_links) - len(days))]
            }
            df = pd.DataFrame(data)
            # df.to_excel('YouTubeVolume.xlsx', index=False)

        print('=========================================================================')
        time.sleep(random.randint(5, 10))

    data = {
        'Link': all_links,
        'Meta Data': meta_data,
        'Day': days,
        'Comments Volume': all_comments_volume
    }
    df = pd.DataFrame(data)
    # df.to_excel('YouTubeVolume.xlsx', index=False)
    return df


def apply_scraping_youtube(search_query):
    driver.get('https://www.youtube.com/results?search_query={}'.format(search_query))
    time.sleep(random.randint(5, 10))

    SCROLL_PAUSE_TIME = 2  # I'll use this variable for code sleeping time
    delay = 30  # delay time for WebDriver (in seconds)
    scrolling = True  # boolean value -> TRUE means that we're still scrolling, FALSE means we're not scrolling anymore
    last_height = driver.execute_script("return document.documentElement.scrollHeight")  # this is our last/current position on the page
    scrolling_attempt = 5  # we'll have 5 attempts before turning scrolling boolean to False

    # video_description = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div#description-inner')))
    while scrolling == True:
        htmlelement = driver.find_element_by_tag_name("body")  # locate html tag
        htmlelement.send_keys(Keys.END)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")  # calculate current position
        time.sleep(SCROLL_PAUSE_TIME)  # make pause because scrolling will take 0.5/1 seconds
        driver.implicitly_wait(30)  # make longer pause if loading new comments takes longer

        if new_height == last_height:  # if current position  is the same as last position, it means we've reached bottom of the page, so we'll break the loop
            scrolling_attempt -= 1
            print(f"scrolling attempt {scrolling_attempt}")
            if (scrolling_attempt == 0):
                scrolling = False  # this will break while loop
        last_height = new_height  # if current position is not the same as last one, we'll set last position as new height
    df = get_youtube_volumes()
    return df


def login_twitter():
    ''' function that logs in to twitter using Hiyam's credentials for now '''

    url = "https://twitter.com/i/flow/login"
    driver.get(url)
    username = WebDriverWait(driver, delay).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[autocomplete="username"]')))
    username.send_keys(TWITTER_USERNAME)
    username.send_keys(Keys.ENTER)

    password = WebDriverWait(driver, delay).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="password"]')))
    password.send_keys(TWITTER_PASSWORD)
    password.send_keys(Keys.ENTER)
    time.sleep(30)


def get_twitter_volumes():
    tweets = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'article[data-testid="tweet"]')))
    links, volumes, days = [], [], []
    for t in tweets:
        try:
            link = t.find_elements_by_css_selector('a')[3].get_attribute('href')
            # comments_volume = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.css-175oi2r.r-1kbdv8c.r-18u37iz.r-1wtj0ep.r-1ye8kvj.r-1s2bzr4 div')))[0].text
            # comments_volume = t.text.split('\n')
            comments_volume = [e.text for e in t.find_elements_by_css_selector('div.css-175oi2r div[role="group"] div.css-175oi2r.r-18u37iz.r-1h0z5md.r-13awgt0')][0]
            if comments_volume.strip() == '':
                comments_volume = '0'
            day = t.find_element_by_css_selector('time').get_attribute('datetime')

            print(link)
            print('Comments volume: {}'.format(comments_volume))
            print('Day: {}'.format(day))
            print('=========================================')

            links.append(link)
            volumes.append(comments_volume)
            days.append(day)

        except:
            pass

        print(len(links), len(volumes), len(days))

    data = {
        'Link': links,
        'Day': days,
        'Comments Volume': volumes
    }
    return data



def apply_scraping_twitter(search_query):
    driver.get('https://twitter.com/search?q={}'.format(search_query))
    time.sleep(random.randint(5, 10))

    SCROLL_PAUSE_TIME = 2  # I'll use this variable for code sleeping time
    delay = 30  # delay time for WebDriver (in seconds)
    scrolling = True  # boolean value -> TRUE means that we're still scrolling, FALSE means we're not scrolling anymore
    last_height = driver.execute_script("return document.documentElement.scrollHeight")  # this is our last/current position on the page
    scrolling_attempt = 5  # we'll have 5 attempts before turning scrolling boolean to False

    all_data = {
        'Link': [],
        'Day': [],
        'Comments Volume': []
    }

    # video_description = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div#description-inner')))
    while scrolling == True:
        htmlelement = driver.find_element_by_tag_name("body")  # locate html tag
        htmlelement.send_keys(Keys.END)

        new_height = driver.execute_script("return document.documentElement.scrollHeight")  # calculate current position
        time.sleep(SCROLL_PAUSE_TIME)  # make pause because scrolling will take 0.5/1 seconds
        driver.implicitly_wait(30)  # make longer pause if loading new comments takes longer

        try:
            data = get_twitter_volumes()
            all_data['Link'].extend(data['Link'])
            all_data['Day'].extend(data['Day'])
            all_data['Comments Volume'].extend(data['Comments Volume'])

        except:
            pass

        if new_height == last_height:  # if current position  is the same as last position, it means we've reached bottom of the page, so we'll break the loop
            scrolling_attempt -= 1
            print(f"scrolling attempt {scrolling_attempt}")
            if (scrolling_attempt == 0):
                scrolling = False  # this will break while loop
        last_height = new_height  # if current position is not the same as last one, we'll set last position as new height

    df = pd.DataFrame(all_data)
    df = df.drop_duplicates()
    # df.to_excel('TwitterVolume.xlsx', index=False)
    return df


def get_facebook_volumes_posts(driver):
    all_data = {'Link': [], 'Day': [], 'scraping_time': [], 'Comments Volume': []}
    all_links_posts = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.x1i10hfl.xjbqb8w.x6umtig.x1b1mbwd.xaqea5y.xav7gou.x9f619.x1ypdohk.xt0psk2.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.xexx8yu.x4uap5.x18d9i69.xkhd6sd.x16tdsg8.x1hl2dhg.xggy1nq.x1a2a7pz.x1heor9g.xt0b8zv.xo1l8bm')))
    all_dates = [a.text for a in all_links_posts]
    for i, post in enumerate(all_links_posts):
        if 'pfbid' in post:
            driver.get(post)
            time.sleep(random.randint(3, 5))
            try:
                comment_nb = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1xmvt09.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.xudqn12.x3x7a5m.x6prxxf.xvq8zen.xo1l8bm.xi81zsa')))[1].text
            except:
                comment_nb = 0
            print('{} - {} comments on the post: {}'.format(all_dates[i], comment_nb, post))

            all_data['Link'].append(post)
            all_data['Day'].append(all_dates[i])
            all_data['scraping_time'].append(str(time.time()))
            all_data['Comments Volume'].append(comment_nb)

    return all_data


def get_facebook_volumes_videos(driver):
    all_data = {'Link': [], 'Day': [], 'scraping_time': [], 'Comments Volume': []}
    all_links_videos = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,  'a.x1i10hfl.xjbqb8w.x6umtig.x1b1mbwd.xaqea5y.xav7gou.x9f619.x1ypdohk.xt0psk2.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.xexx8yu.x4uap5.x18d9i69.xkhd6sd.x16tdsg8.x1hl2dhg.xggy1nq.x1a2a7pz.x1heor9g.xt0b8zv.x1s688f')))
    all_links_videos = [a.get_attribute('href') for a in all_links_videos]

    for watch in all_links_videos:
        driver.get(watch)
        time.sleep(random.randint(3, 5))
        comments_watch = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1xmvt09.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x4zkp8e.x676frb.x1nxh6w3.x1sibtaa.xo1l8bm.xi81zsa')))[0].text
        date_watch = WebDriverWait(driver, delay).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.x1i10hfl.xjbqb8w.x6umtig.x1b1mbwd.xaqea5y.xav7gou.x9f619.x1ypdohk.xt0psk2.xe8uvvx.xdj266r.x11i5rnm.xat24cr.x1mh8g0r.xexx8yu.x4uap5.x18d9i69.xkhd6sd.x16tdsg8.x1hl2dhg.xggy1nq.x1a2a7pz.x1heor9g.xt0b8zv.xo1l8bm')))[0].text
        print('{} - {} comments on the post: {}'.format(date_watch, comments_watch, watch))

        all_data['Link'].append(watch)
        all_data['Day'].append(date_watch)
        all_data['scraping_time'].append(str(time.time()))
        all_data['Comments Volume'].append(comments_watch)

    return all_data


# def add_volume_sheet_youtube(save_dir, df1, time_after, time_before):
#     file_name = 'E'
#     if time_after is not None:
#         file_name += time_after
#     if time_before is not None:
#         file_name += '-' + time_before
#     file_name += '-DB.txt'
#     db_links = list(df['Link'])
#     with open(os.path.join(save_dir, file_name), 'w') as f:
#         for link in db_links:
#             f.write(link + '\n')
#     f.close()
#
#     writer = pd.ExcelWriter(os.path.join(save_dir, 'YouTubeVolumeMod.xlsx'), engine='xlsxwriter')
#     df2 = df1
#     comments_volume_updated = []
#     for l in list(df1['Comments Volume']):
#         try:
#             comments_volume_updated.append(int(l.split(' ')[0]))
#         except:
#             comments_volume_updated.append(0)
#     df2['Comments Volume'] = comments_volume_updated
#
#     days_unique = list(set(list(df1['Day'])))
#     data = []
#     for d in days_unique:
#         df_sub = df2[df2['Day'] == d]
#         volume = 0
#         links = len(df_sub)
#         for i, row in df_sub.iterrows():
#             volume += int(row['Comments Volume'])
#
#         data.append([d, links, volume])
#
#     df3 = pd.DataFrame(data, columns=['Day', 'Links', 'Comments Volume'])
#
#     df1.to_excel(writer, sheet_name='original', index=False)
#     df3.to_excel(writer, sheet_name='volume aggregation', index=False)
#     writer.close()
#
#
# def add_volume_sheet_twitter(save_dir, df1, time_after, time_before):
#     file_name = 'E'
#     if time_after is not None:
#         file_name += time_after
#     if time_before is not None:
#         file_name += '-' + time_before
#     file_name += '-DB.txt'
#     db_links = list(df['Link'])
#     with open(os.path.join(save_dir, file_name), 'w') as f:
#         for link in db_links:
#             f.write(link + '\n')
#     f.close()
#
#     writer = pd.ExcelWriter(os.path.join(save_dir, 'TwitterVolumeMod.xlsx'), engine='xlsxwriter')
#     df2 = df1
#     dates_updated = [l[:10] for l in list(df1['Day'])]
#     df2['Day'] = dates_updated
#
#     days_unique = list(set(list(df2['Day'])))
#     data = []
#     for d in days_unique:
#         df_sub = df2[df2['Day'] == d]
#         volume = 0
#         links = len(df_sub)
#         for i, row in df_sub.iterrows():
#             try:
#                 volume += int(row['Comments Volume'])
#             except:
#                 volume += 0
#
#         data.append([d, links, volume])
#
#     df3 = pd.DataFrame(data, columns=['Day', 'Links', 'Comments Volume'])
#
#     df1.to_excel(writer, sheet_name='original', index=False)
#     df3.to_excel(writer, sheet_name='volume aggregation', index=False)
#
#     writer.close()


if __name__ == '__main__':

    TWITTER_USERNAME = input("Please enter your Twitter Username: ")
    TWITTER_PASSWORD = input("Please enter your Twitter Password: ")

    category2english = {
        'فساد': 'corruption',
        'جندري': 'gender',
        'جنساني': 'sexuality',
        'ديني': 'religion',
        'سياسة وأمن': 'politics and security',
        'لجوء': 'refugee'
    }


    driver = webdriver.Chrome('C:\Program Files\chromedriver-win32-119\chromedriver.exe')
    delay = 30
    events_file = input("Please enter the name of the events file: ")
    df = pd.read_excel(events_file.strip())

    months = input("Please enter the starting and ending month (inclusive) as MM-MM: ")
    months = months.strip().split("-")
    months_nb = []
    for m in months:
        if '0' in m:
            months_nb.append(int(m[1]))
        else:
            months_nb.append(m)

    months_nb = list(range(months_nb[0], months_nb[1]+1))

    print('############################################# YOUTUBE ##############################################################################')
    for i, row in tqdm(df.iterrows(), total=df.shape[0]):

        time_after = row['التاريخ']

        if '2023' not in str(time_after): # restricting the years 2023
            continue

        time_after_dt = datetime.strptime(str(time_after)[:10], '%Y-%m-%d')
        month_curr = time_after_dt.month

        if month_curr not in months_nb:
            continue

        category = str(row['النوع'])
        category_en = category2english[category]  # get the english equivalent of the category

        window = str(row['Window of Date'])

        if window.strip() in ['', 'nan']:
            continue

        if 'week' in window:
            num = window.split(' ')[0]
            # time_before_dt = relativedelta(weeks=int(num)) + time_after_dt
            # time_before_dt = time_after_dt + relativedelta(weeks=1)
        else:
            num = window.split(' ')[0]
            # time_before_dt = relativedelta(months=int(num)) + time_after_dt
            # time_before_dt = time_after_dt + relativedelta(weeks=1)

        time_before_dt = relativedelta(weeks=2) + time_after_dt

        keywords_text = str(row['كلمات مفتاحية']).strip()
        keywords_text_facebook = str(row['كلمات مفتاحية']).strip()

        if keywords_text in ["", "nan"]:
            continue

        print('Keywords: {}'.format(keywords_text))
        print('start date: {}'.format(str(time_after_dt)))
        print('end date: {}'.format(str(time_before_dt)))

        keywords_text_new = line_split(keywords_text)
        j = 0
        for kt in keywords_text_new:
            ktnew = []
            if kt == '':
                continue

            kt = kt.replace('\'', '')
            kt = kt.replace('\"', '')
            kt = kt.strip()

            for i, k in enumerate(kt.split(" ")):
                if i == 0:
                    if len(kt.split(" ")) > 1:
                        knew = "\"" + k + "\" AND"
                    else:
                        knew = "\"" + k + "\""
                elif i == len(kt.split(" ")) - 1:
                    knew = "\"" + k + "\""
                else:
                    knew = "\"" + k + "\" AND"

                ktnew.append(knew)

            kt = " ".join(ktnew)

            youtube_queries = []

            ktyoutube = "intitle:{}".format(kt)
            # kt = ' AND '.join(kt.split(' '))

            save_dir_volume = 'Volume-New/{}/Total-KD/'.format(category_en)
            save_dir_links = 'Events-New/E{}-KD/'.format(str(time_after_dt)[:10])

            mkdir(save_dir_volume)
            mkdir(save_dir_links)
            #
            # writer_youtube = pd.ExcelWriter(os.path.join(save_dir_volume, 'YouTube.xlsx'), engine='xlsxwriter')
            # writer_twitter = pd.ExcelWriter(os.path.join(save_dir_volume, 'Twitter.xlsx'), engine='xlsxwriter')

            query_youtube = ''
            query_twitter = ''

            query_youtube += ktyoutube.strip()
            query_youtube += '+after%3A{}'.format(str(time_after_dt)[:10])
            query_youtube += '+before%3A{}'.format(str(time_before_dt)[:10])

            youtube_queries.append(query_youtube)

            query_youtube2 = ''
            ktyoutube = kt
            query_youtube2 += ktyoutube.strip()
            query_youtube2 += '+after%3A{}'.format(str(time_after_dt)[:10])
            query_youtube2 += '+before%3A{}'.format(str(time_before_dt)[:10])

            youtube_queries.append(query_youtube2)

            ######################################## YOUTUBE ############################################

            for yq in youtube_queries:

                try:
                    t1 = time.time()
                    df = apply_scraping_youtube(search_query=yq)
                    if not os.path.isfile(os.path.join(save_dir_volume, 'YouTube.xlsx')):
                        with pd.ExcelWriter(os.path.join(save_dir_volume, 'YouTube.xlsx')) as writer:
                            df.to_excel(writer, sheet_name='E{}_{}'.format(str(time_after_dt)[:10], j), index=False)
                        writer.close()
                    else:
                        book = load_workbook(os.path.join(save_dir_volume, 'YouTube.xlsx'))
                        writer = pd.ExcelWriter(os.path.join(save_dir_volume, 'YouTube.xlsx'), engine='openpyxl')
                        writer.book = book

                        df.to_excel(writer, sheet_name='E{}-{}'.format(str(time_after_dt)[:10], j), index=False)
                        writer.close()

                    t2 = time.time()
                    print('This process took {} mins'.format((t2 - t1) / 60))

                    file_name = 'E{}-DB.txt'.format(str(time_after_dt)[:10])
                    db_links = list(df['Link'])
                    if not os.path.isfile(os.path.join(save_dir_links, file_name)):
                        with open(os.path.join(save_dir_links, file_name), 'w', encoding='utf-8') as f:
                            try:
                                f.write('{}\n'.format(yq))
                            except:
                                f.write('{}\n'.format(yq.encode('utf-8')))
                            for link in db_links:
                                f.write(link + '\n')
                        f.close()
                    else:
                        with open(os.path.join(save_dir_links, file_name), 'a', encoding='utf-8') as f:
                            try:
                                f.write('{}\n'.format(yq))
                            except:
                                f.write('{}\n'.format(yq.encode('utf-8')))
                            for link in db_links:
                                f.write(link + '\n')
                        f.close()

                except:
                    print('Did not find results for query {}'.format(keywords_text))

                j += 1


    # ############################################# TWITTER ##############################################################################
    print('############################################# TWITTER ##############################################################################')
    login_twitter() # log in once

    for i, row in tqdm(df.iterrows(), total=df.shape[0]):
        time_after = row['التاريخ']

        if '2023' not in str(time_after):  # restricting the years 2023
            continue

        time_after_dt = datetime.strptime(str(time_after)[:10], '%Y-%m-%d')
        month_curr = time_after_dt.month

        if month_curr not in months_nb:
            continue

        category = str(row['النوع'])
        category_en = category2english[category]  # get the english equivalent of the category

        window = str(row['Window of Date'])

        if window.strip() in ['', 'nan']:
            continue

        if 'week' in window:
            num = window.split(' ')[0]
            # time_before_dt = relativedelta(weeks=int(num)) + time_after_dt
            # time_before_dt = time_after_dt + relativedelta(weeks=1)
        else:
            num = window.split(' ')[0]
            # time_before_dt = relativedelta(months=int(num)) + time_after_dt
            # time_before_dt = time_after_dt + relativedelta(weeks=1)

        time_before_dt = relativedelta(weeks=2) + time_after_dt

        keywords_text = str(row['كلمات مفتاحية']).strip()
        keywords_text_facebook = str(row['كلمات مفتاحية']).strip()

        if keywords_text in ["", "nan"]:
            continue

        print('Keywords: {}'.format(keywords_text))
        print('start date: {}'.format(str(time_after_dt)))
        print('end date: {}'.format(str(time_before_dt)))

        keywords_text_new = line_split(keywords_text)
        j = 0
        for kt in keywords_text_new:
            ktnew = []
            if kt == '':
                continue

            kt = kt.replace('\'', '')
            kt = kt.replace('\"', '')
            kt = kt.strip()

            for i, k in enumerate(kt.split(" ")):
                if i == 0:
                    if len(kt.split(" ")) > 1:
                        knew = "\"" + k + "\" AND"
                    else:
                        knew = "\"" + k + "\""
                elif i == len(kt.split(" ")) - 1:
                    knew = "\"" + k + "\""
                else:
                    knew = "\"" + k + "\" AND"

                ktnew.append(knew)

            kt = " ".join(ktnew)

            ktyoutube = "intitle:{}".format(kt)
            # kt = ' AND '.join(kt.split(' '))

            save_dir_volume = 'Volume-New/{}/Total-KD/'.format(category_en)
            save_dir_links = 'Events-New/E{}-KD/'.format(str(time_after_dt)[:10])

            mkdir(save_dir_volume)
            mkdir(save_dir_links)

            query_youtube = ''
            query_twitter = ''

            ######################################## TWITTER ############################################
            query_twitter += kt.strip()
            query_twitter += 'until%3A{}'.format(str(time_before_dt)[:10])
            query_twitter += '%20since%3A{}'.format(str(time_after_dt)[:10])
            query_twitter += '&src=typed_query'

            query_twitter2 = query_twitter + '&f=live'

            twitter_queries = []
            twitter_queries.append(query_twitter)
            twitter_queries.append(query_twitter2)

            for qt in twitter_queries:
                t1 = time.time()
                df = apply_scraping_twitter(search_query=qt)
                if not os.path.isfile(os.path.join(save_dir_volume, 'Twitter.xlsx')):
                    with pd.ExcelWriter(os.path.join(save_dir_volume, 'Twitter.xlsx')) as writer:
                        df.to_excel(writer, sheet_name='E{}'.format(str(time_after_dt)[:10]), index=False)
                    writer.close()

                else:
                    book = load_workbook(os.path.join(save_dir_volume, 'Twitter.xlsx'))
                    writer = pd.ExcelWriter(os.path.join(save_dir_volume, 'Twitter.xlsx'), engine='openpyxl')
                    writer.book = book
                    df.to_excel(writer, sheet_name='E{}-{}'.format(str(time_after_dt)[:10], j), index=False)
                    writer.close()

                t2 = time.time()

                file_name = 'E{}-Twitter.txt'.format(str(time_after_dt)[:10])
                db_links = list(df['Link'])
                save_dir_links_twitter = save_dir_links + 'E{}-Twitter/'.format(str(time_after_dt)[:10])
                mkdir(save_dir_links_twitter)
                if not os.path.isfile(os.path.join(save_dir_links_twitter, file_name)):
                    with open(os.path.join(save_dir_links_twitter, file_name), 'w', encoding='utf-8') as f:
                        # f.write('{}\n'.format(query_twitter))
                        try:
                            f.write('{}\n'.format(qt))
                        except:
                            f.write('{}\n'.format(qt.encode('utf-8')))
                        for link in db_links:
                            f.write(link + '\n')
                    f.close()
                else:
                    with open(os.path.join(save_dir_links_twitter, file_name), 'a', encoding='utf-8') as f:
                        try:
                            f.write('{}\n'.format(qt))
                        except:
                            f.write('{}\n'.format(qt.encode('utf-8')))
                        for link in db_links:
                            f.write(link + '\n')
                    f.close()

                print('This process took {} mins'.format((t2 - t1) / 60))

                j += 1

            time.sleep(random.randint(120, 180))






